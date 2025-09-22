
"""Vehicle inventory management panel for Logistics module.

This widget implements a desktop-focused split view combining a searchable
and filterable table of vehicles with a contextual detail pane.  It follows
the specification provided in the task description including pagination,
status pills, tag chips, keyboard shortcuts, an import wizard, and an export
modal that runs in the background.
"""

from __future__ import annotations

import csv
import math
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Optional, Sequence

from PySide6 import QtConcurrent
from PySide6.QtCore import (
    QAbstractTableModel,
    QEvent,
    QItemSelectionModel,
    QModelIndex,
    QObject,
    QPoint,
    QRect,
    QSize,
    Qt,
    QTimer,
    QSettings,
    Signal,
    QThread,
)

try:  # QtConcurrent extras were removed from some builds
    from PySide6.QtCore import QFutureWatcher
except ImportError:  # pragma: no cover - depends on PySide6 build
    QFutureWatcher = None  # type: ignore[assignment]
from PySide6.QtGui import (
    QAction,
    QColor,
    QKeySequence,
    QPainter,
    QPalette,
    QPen,
    QShortcut,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLayout,
    QLayoutItem,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QStackedLayout,
    QStyle,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QWizard,
    QWizardPage,
    QProgressBar,
)

from notifications.models import Notification
from notifications.services import get_notifier

from modules.logistics.vehicle.panels.vehicle_edit_window import VehicleEditDialog, VehicleRepository

__all__ = ["VehicleInventoryPanel"]


class _ExportWorkerThread(QThread):
    """Run export tasks in a thread when QFutureWatcher is unavailable."""

    completed = Signal(object)
    failed = Signal(str)

    def __init__(
        self,
        task: Callable[[dict[str, Any]], dict[str, Any]],
        params: dict[str, Any],
        parent: QObject | None = None,
    ) -> None:
        super().__init__(parent)
        self._task = task
        self._params = params

    def run(self) -> None:  # type: ignore[override]
        try:
            result = self._task(self._params)
        except Exception as exc:  # pragma: no cover - depends on runtime state
            self.failed.emit(str(exc))
        else:
            self.completed.emit(result)


# ---------------------------------------------------------------------------
# Helper data structures
# ---------------------------------------------------------------------------


STATUS_COLORS: dict[str, tuple[str, str]] = {
    "available": ("#2e7d32", "#ffffff"),
    "in service": ("#1565c0", "#ffffff"),
    "out of service": ("#757575", "#ffffff"),
    "retired": ("#424242", "#ffffff"),
}

DEFAULT_TAG_BG = "#eceff1"
DEFAULT_TAG_FG = "#37474f"


def middle_ellipsis(text: str, max_chars: int = 18) -> str:
    """Collapse the middle of ``text`` so it fits within ``max_chars``."""

    if not text:
        return ""
    if len(text) <= max_chars:
        return text
    half = max_chars // 2 - 1
    return f"{text[:half]}…{text[-half:]}"


def format_timestamp(value: Any) -> str | None:
    """Return a human-friendly ``YYYY-MM-DD HH:MM`` timestamp string."""

    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value).strftime("%Y-%m-%d %H:%M")
    text = str(value).strip()
    if not text:
        return None

    # Common ISO variants
    cleaned = text.replace("Z", "+00:00")
    for fmt in (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f%z",
    ):
        try:
            dt = datetime.strptime(cleaned, fmt)
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return text


def coalesce(values: Iterable[Any]) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def ensure_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str):
        return [part.strip() for part in value.split(",") if part.strip()]
    return []


@dataclass
class ColumnDefinition:
    key: str
    title: str
    sort_key: str
    alignment: Qt.AlignmentFlag = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft


COLUMNS: tuple[ColumnDefinition, ...] = (
    ColumnDefinition("id", "ID", "id"),
    ColumnDefinition("license_plate", "License Plate", "license_plate"),
    ColumnDefinition("vin", "VIN", "vin"),
    ColumnDefinition("vehicle", "Vehicle (Year Make Model)", "vehicle"),
    ColumnDefinition("capacity", "Cap", "capacity", Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
    ColumnDefinition("type", "Type", "type"),
    ColumnDefinition("status", "Status", "status", Qt.AlignmentFlag.AlignCenter),
    ColumnDefinition("tags", "Tags", "tags"),
)


TAG_ROLE = Qt.ItemDataRole.UserRole + 1
RECORD_ROLE = Qt.ItemDataRole.UserRole + 2


@dataclass
class VehicleRecord:
    identifier: str
    license_plate: str | None
    vin: str | None
    year: int | None
    make: str | None
    model: str | None
    capacity: int | None
    type_id: str | None
    type_label: str | None
    status_id: str | None
    status_label: str | None
    tags: list[str] = field(default_factory=list)
    created: str | None = None
    updated: str | None = None
    organization: str | None = None
    raw: dict[str, Any] = field(default_factory=dict)

    def vehicle_label(self) -> str:
        parts: list[str] = []
        if self.year:
            parts.append(str(self.year))
        if self.make:
            parts.append(str(self.make))
        if self.model:
            parts.append(str(self.model))
        if not parts:
            return "(none)"
        return " ".join(parts)

    def tags_display(self) -> str:
        return ", ".join(self.tags)


class VehicleInventoryModel(QAbstractTableModel):
    """Table model presenting vehicle inventory records."""

    def __init__(self, parent: QObject | None = None) -> None:
        super().__init__(parent)
        self._records: list[VehicleRecord] = []

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]
        return 0 if parent.isValid() else len(self._records)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # type: ignore[override]
        return 0 if parent.isValid() else len(COLUMNS)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> Any:  # type: ignore[override]
        if not index.isValid() or not (0 <= index.row() < len(self._records)):
            return None
        record = self._records[index.row()]
        column = COLUMNS[index.column()]

        if role == Qt.ItemDataRole.DisplayRole:
            if column.key == "id":
                return record.identifier
            if column.key == "license_plate":
                return record.license_plate or "—"
            if column.key == "vin":
                return middle_ellipsis(record.vin or "") or "—"
            if column.key == "vehicle":
                return record.vehicle_label()
            if column.key == "capacity":
                return str(record.capacity) if record.capacity not in (None, "") else "—"
            if column.key == "type":
                return record.type_label or "—"
            if column.key == "status":
                return record.status_label or "—"
            if column.key == "tags":
                return record.tags_display()

        if role == Qt.ItemDataRole.ToolTipRole:
            if column.key == "vin" and record.vin:
                return record.vin
            if column.key == "tags" and record.tags:
                return record.tags_display()
            if column.key == "license_plate" and record.license_plate:
                return record.license_plate

        if role == TAG_ROLE and column.key == "tags":
            return record.tags

        if role == RECORD_ROLE:
            return record

        if role == Qt.ItemDataRole.TextAlignmentRole:
            return int(column.alignment)

        if role == Qt.ItemDataRole.AccessibleTextRole and column.key == "status":
            return f"Status {record.status_label}" if record.status_label else "Status unknown"

        return None

    def headerData(
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.ItemDataRole.DisplayRole,
    ) -> Any:  # type: ignore[override]
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return COLUMNS[section].title
        return super().headerData(section, orientation, role)

    def flags(self, index: QModelIndex) -> Qt.ItemFlags:  # type: ignore[override]
        if not index.isValid():
            return Qt.ItemFlag.ItemIsEnabled
        return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable

    def set_records(self, records: Sequence[VehicleRecord]) -> None:
        self.beginResetModel()
        self._records = list(records)
        self.endResetModel()

    def record(self, row: int) -> VehicleRecord | None:
        if 0 <= row < len(self._records):
            return self._records[row]
        return None

    def sort_key_for_column(self, column: int) -> str:
        return COLUMNS[column].sort_key


class StatusPillDelegate(QStyledItemDelegate):
    """Paint status values as rounded pills with high contrast."""

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex) -> None:  # type: ignore[override]
        status_text = str(index.data(Qt.ItemDataRole.DisplayRole) or "")
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)

        base = status_text.lower()
        bg_color, fg_color = STATUS_COLORS.get(base, ("#5f6368", "#ffffff"))
        rect = option.rect.adjusted(6, 8, -6, -8)
        radius = rect.height() / 2

        color = QColor(bg_color)
        painter.setBrush(color)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(rect, radius, radius)

        painter.setPen(QPen(QColor(fg_color)))
        font = option.font
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, status_text or "—")
        painter.restore()

    def sizeHint(self, option: QStyleOptionViewItem, index: QModelIndex) -> QSize:  # type: ignore[override]
        text = str(index.data(Qt.ItemDataRole.DisplayRole) or "") or "—"
        metrics = option.fontMetrics
        width = metrics.horizontalAdvance(text) + 24
        height = max(metrics.height() + 16, 28)
        return QSize(width, height)


class TagsChipDelegate(QStyledItemDelegate):
    """Render tags as wrapping chips with overflow indicator."""

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex) -> None:  # type: ignore[override]
        tags: list[str] = index.data(TAG_ROLE) or []
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        rect = option.rect.adjusted(0, 10, 0, -10)

        metrics = option.fontMetrics
        x = rect.left() + 4
        available_width = rect.width()

        chips_drawn = 0
        hidden = 0

        for tag in tags:
            chip_width = metrics.horizontalAdvance(tag) + 20
            if x + chip_width > rect.left() + available_width:
                hidden += 1
                continue

            chip_rect = QRect(x, rect.top(), chip_width, rect.height())
            self._draw_chip(painter, chip_rect, tag)
            x += chip_width + 6
            chips_drawn += 1

        if hidden > 0:
            more_label = f"+{hidden}"
            chip_width = metrics.horizontalAdvance(more_label) + 20
            chip_rect = QRect(x, rect.top(), chip_width, rect.height())
            self._draw_chip(painter, chip_rect, more_label)

        if chips_drawn == 0 and hidden == 0:
            painter.setPen(QPen(option.palette.color(QPalette.ColorRole.Disabled, QPalette.ColorRole.Text)))
            painter.drawText(rect, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, "—")

        painter.restore()

    def _draw_chip(self, painter: QPainter, rect: QRect, text: str) -> None:
        radius = rect.height() / 2
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(DEFAULT_TAG_BG))
        painter.drawRoundedRect(rect, radius, radius)
        painter.setPen(QColor(DEFAULT_TAG_FG))
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, text)

    def sizeHint(self, option: QStyleOptionViewItem, index: QModelIndex) -> QSize:  # type: ignore[override]
        metrics = option.fontMetrics
        return QSize(option.rect.width(), max(metrics.height() + 16, 28))


class FlowLayout(QLayout):
    """Simple flow layout that wraps child widgets."""

    def __init__(self, parent: QWidget | None = None, margin: int = 0, spacing: int = 6) -> None:
        super().__init__(parent)
        self.setContentsMargins(margin, margin, margin, margin)
        self._spacing = spacing
        self._items: list[QLayoutItem] = []

    def addItem(self, item: QLayoutItem) -> None:  # type: ignore[override]
        self._items.append(item)

    def count(self) -> int:  # type: ignore[override]
        return len(self._items)

    def itemAt(self, index: int) -> QLayoutItem | None:  # type: ignore[override]
        if 0 <= index < len(self._items):
            return self._items[index]
        return None

    def takeAt(self, index: int) -> QLayoutItem | None:  # type: ignore[override]
        if 0 <= index < len(self._items):
            return self._items.pop(index)
        return None

    def expandingDirections(self) -> Qt.Orientations:  # type: ignore[override]
        return Qt.Orientation(0)

    def hasHeightForWidth(self) -> bool:  # type: ignore[override]
        return True

    def heightForWidth(self, width: int) -> int:  # type: ignore[override]
        return self._do_layout(QRect(0, 0, width, 0), test_only=True)

    def setGeometry(self, rect: QRect) -> None:  # type: ignore[override]
        super().setGeometry(rect)
        self._do_layout(rect, test_only=False)

    def sizeHint(self) -> QSize:  # type: ignore[override]
        return QSize()

    def _do_layout(self, rect: QRect, test_only: bool) -> int:
        x = rect.x()
        y = rect.y()
        line_height = 0

        for item in self._items:
            wid = item.widget()
            if wid is not None and not wid.isVisible():
                continue
            item_size = item.sizeHint()
            if item_size.width() > rect.width() and x == rect.x():
                item_size.setWidth(rect.width())
            next_x = x + item_size.width() + self._spacing
            if next_x - self._spacing > rect.right() and line_height > 0:
                x = rect.x()
                y = y + line_height + self._spacing
                next_x = x + item_size.width() + self._spacing
                line_height = item_size.height()
            else:
                line_height = max(line_height, item_size.height())

            if not test_only:
                item.setGeometry(QRect(QPoint(x, y), item_size))
            x = next_x

        return y + line_height - rect.y()


class VehicleDetailPanel(QFrame):
    """Detail panel that renders the selected vehicle's information."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("vehicleDetailPanel")
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(
            """
            #vehicleDetailPanel {
                border: 1px solid palette(Midlight);
                border-radius: 12px;
                background: palette(Base);
            }
            """
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        title = QLabel("Detail")
        title.setStyleSheet("font-size: 18px; font-weight: 600;")
        layout.addWidget(title)

        self._stack = QStackedLayout()
        layout.addLayout(self._stack)

        placeholder = QLabel("Select a vehicle to view details.")
        placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder.setStyleSheet("color: palette(Mid);")
        placeholder_wrap = QWidget()
        placeholder_layout = QVBoxLayout(placeholder_wrap)
        placeholder_layout.addStretch(1)
        placeholder_layout.addWidget(placeholder)
        placeholder_layout.addStretch(1)
        self._stack.addWidget(placeholder_wrap)

        self._detail_widget = QWidget()
        detail_layout = QGridLayout(self._detail_widget)
        detail_layout.setHorizontalSpacing(12)
        detail_layout.setVerticalSpacing(8)
        detail_layout.setColumnStretch(1, 1)
        self._stack.addWidget(self._detail_widget)

        self._fields: dict[str, QLabel] = {}
        self._status_chip = QLabel()
        self._status_chip.setObjectName("vehicleStatusChip")
        self._status_chip.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._status_chip.setMinimumHeight(28)
        self._status_chip.setStyleSheet(
            """
            #vehicleStatusChip {
                border-radius: 14px;
                padding: 4px 12px;
                font-weight: 600;
            }
            """
        )

        self._tags_container = QWidget()
        self._tags_layout = FlowLayout(self._tags_container, margin=0, spacing=6)

        fields = [
            ("Vehicle ID", "identifier"),
            ("License Plate", "license_plate"),
            ("VIN", "vin"),
            ("Vehicle", "vehicle"),
            ("Capacity", "capacity"),
            ("Type", "type"),
            ("Status", "status"),
            ("Tags", "tags"),
            ("Created", "created"),
            ("Updated", "updated"),
        ]

        for row, (label_text, key) in enumerate(fields):
            label = QLabel(label_text + ":")
            label.setStyleSheet("font-weight: 500;")
            detail_layout.addWidget(label, row, 0)

            if key == "status":
                detail_layout.addWidget(self._status_chip, row, 1)
                continue
            if key == "tags":
                detail_layout.addWidget(self._tags_container, row, 1)
                continue

            value_label = QLabel("")
            value_label.setObjectName(f"detail_{key}")
            if key in {"license_plate", "vin"}:
                mono = value_label.font()
                mono.setFamily("Monospace")
                value_label.setFont(mono)
            value_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            detail_layout.addWidget(value_label, row, 1)
            self._fields[key] = value_label

        layout.addStretch(1)

    def set_record(self, record: VehicleRecord | None) -> None:
        if record is None:
            self._stack.setCurrentIndex(0)
            return

        self._stack.setCurrentIndex(1)
        self._fields["identifier"].setText(str(record.identifier))
        self._fields["license_plate"].setText(record.license_plate or "—")
        self._fields["vin"].setText(record.vin or "—")
        self._fields["vehicle"].setText(record.vehicle_label())
        self._fields["capacity"].setText(str(record.capacity) if record.capacity not in (None, "") else "—")
        self._fields["type"].setText(record.type_label or "—")

        status = record.status_label or "—"
        self._status_chip.setText(status)
        status_colors = STATUS_COLORS.get((record.status_label or "").lower(), ("#5f6368", "#ffffff"))
        self._status_chip.setStyleSheet(
            """
            #vehicleStatusChip {
                border-radius: 14px;
                padding: 4px 12px;
                font-weight: 600;
                background-color: %s;
                color: %s;
            }
            """
            % status_colors
        )

        while self._tags_layout.count():
            item = self._tags_layout.takeAt(0)
            if item:
                widget = item.widget()
                if widget:
                    widget.deleteLater()

        if record.tags:
            for tag in record.tags:
                chip = QLabel(tag)
                chip.setStyleSheet(
                    "border-radius: 12px; background: %s; color: %s; padding: 4px 10px;"
                    % (DEFAULT_TAG_BG, DEFAULT_TAG_FG)
                )
                chip.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                self._tags_layout.addWidget(chip)
        else:
            chip = QLabel("—")
            chip.setStyleSheet("color: palette(Mid);")
            self._tags_layout.addWidget(chip)

        self._fields["created"].setText(record.created or "—")
        self._fields["updated"].setText(record.updated or "—")


class VehicleFilterBar(QWidget):
    """Search and filter controls for the vehicle inventory table."""

    filtersChanged = Signal()

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search vehicles…")
        self.search_edit.setClearButtonEnabled(True)
        search_action = QAction(self)
        search_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogContentsView))
        self.search_edit.addAction(search_action, QLineEdit.ActionPosition.LeadingPosition)
        layout.addWidget(self.search_edit, stretch=2)

        self.type_combo = QComboBox()
        self.type_combo.setAccessibleName("Vehicle type filter")
        layout.addWidget(self.type_combo)

        self.status_combo = QComboBox()
        self.status_combo.setAccessibleName("Vehicle status filter")
        layout.addWidget(self.status_combo)

        self.reset_button = QToolButton()
        self.reset_button.setText("Reset filters")
        self.reset_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.reset_button.setStyleSheet("QToolButton { text-decoration: underline; border: none; padding: 4px; }")
        self.reset_button.hide()
        layout.addWidget(self.reset_button)
        layout.addStretch(1)

        self._debounce = QTimer(self)
        self._debounce.setInterval(250)
        self._debounce.setSingleShot(True)
        self._debounce.timeout.connect(self._emit_filters)

        self.search_edit.textChanged.connect(self._on_search_text)
        self.type_combo.currentIndexChanged.connect(self._emit_filters)
        self.status_combo.currentIndexChanged.connect(self._emit_filters)
        self.reset_button.clicked.connect(self.reset_filters)

        self._type_options: list[dict[str, Any]] = []
        self._status_options: list[dict[str, Any]] = []

    def focus_search(self) -> None:
        self.search_edit.setFocus(Qt.FocusReason.ShortcutFocusReason)

    def apply_reference_options(
        self,
        types: Sequence[dict[str, Any]],
        statuses: Sequence[dict[str, Any]],
    ) -> None:
        self._type_options = list(types)
        self._status_options = list(statuses)

        current_type = self.type_combo.currentText()
        current_status = self.status_combo.currentText()

        self.type_combo.blockSignals(True)
        self.status_combo.blockSignals(True)

        self.type_combo.clear()
        self.type_combo.addItem("All", None)
        for entry in types:
            self.type_combo.addItem(str(entry.get("name")), entry.get("id"))

        self.status_combo.clear()
        self.status_combo.addItem("All", None)
        for entry in statuses:
            self.status_combo.addItem(str(entry.get("name")), entry.get("id"))

        if current_type:
            index = self.type_combo.findText(current_type)
            if index >= 0:
                self.type_combo.setCurrentIndex(index)
        if current_status:
            index = self.status_combo.findText(current_status)
            if index >= 0:
                self.status_combo.setCurrentIndex(index)

        self.type_combo.blockSignals(False)
        self.status_combo.blockSignals(False)

    def filters(self) -> dict[str, Any]:
        return {
            "search": self.search_edit.text().strip(),
            "type": self.type_combo.currentData(),
            "status": self.status_combo.currentData(),
        }

    def reset_filters(self) -> None:
        self.search_edit.clear()
        self.type_combo.setCurrentIndex(0)
        self.status_combo.setCurrentIndex(0)
        self._emit_filters()

    def _on_search_text(self, _: str) -> None:
        self._debounce.start()
        self._update_reset_visibility()

    def _emit_filters(self) -> None:
        self._debounce.stop()
        self._update_reset_visibility()
        self.filtersChanged.emit()

    def _update_reset_visibility(self) -> None:
        active = bool(self.search_edit.text().strip()) or self.type_combo.currentIndex() > 0 or self.status_combo.currentIndex() > 0
        self.reset_button.setVisible(active)


class PaginationControls(QWidget):
    """Pagination footer with status text and navigation buttons."""

    pageRequested = Signal(int)
    pageSizeChanged = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.status_label = QLabel("0–0 of 0")
        layout.addWidget(self.status_label)

        layout.addStretch(1)

        self.page_size_combo = QComboBox()
        for size in (20, 50, 100):
            self.page_size_combo.addItem(f"{size} / page", size)
        layout.addWidget(self.page_size_combo)

        self.prev_button = QToolButton()
        self.prev_button.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack))
        layout.addWidget(self.prev_button)

        self.next_button = QToolButton()
        self.next_button.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowForward))
        layout.addWidget(self.next_button)

        self.prev_button.clicked.connect(lambda: self.pageRequested.emit(self._current_page - 1))
        self.next_button.clicked.connect(lambda: self.pageRequested.emit(self._current_page + 1))
        self.page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)

        self._current_page = 1
        self._total_pages = 1
        self._page_size = 20

    def update_state(self, *, total: int, page: int, page_size: int) -> None:
        self._current_page = max(1, page)
        self._page_size = page_size
        self.page_size_combo.blockSignals(True)
        idx = self.page_size_combo.findData(page_size)
        if idx >= 0:
            self.page_size_combo.setCurrentIndex(idx)
        self.page_size_combo.blockSignals(False)

        if total == 0:
            self.status_label.setText("0 of 0")
            self.prev_button.setEnabled(False)
            self.next_button.setEnabled(False)
            return

        total_pages = math.ceil(total / page_size) if page_size else 1
        self._total_pages = max(1, total_pages)
        page = min(max(1, page), self._total_pages)
        start = (page - 1) * page_size + 1
        end = min(total, page * page_size)
        self.status_label.setText(f"{start}–{end} of {total}")
        self.prev_button.setEnabled(page > 1)
        self.next_button.setEnabled(page < self._total_pages)

    def _on_page_size_changed(self) -> None:
        size = self.page_size_combo.currentData()
        if size:
            self.pageSizeChanged.emit(int(size))


class VehicleExportDialog(QDialog):
    """Modal dialog used to configure vehicle export options."""

    def __init__(
        self,
        *,
        parent: QWidget | None = None,
        allow_selected: bool = True,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Export Vehicles")
        self.setModal(True)
        self.setMinimumWidth(420)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        scope_label = QLabel("Scope:")
        scope_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(scope_label)

        self.scope_combo = QComboBox()
        self.scope_combo.addItem("All vehicles", "all")
        self.scope_combo.addItem("Current filters", "filters")
        self.scope_combo.addItem("Selected rows only", "selected")
        if not allow_selected:
            index = self.scope_combo.findData("selected")
            if index >= 0:
                self.scope_combo.model().item(index).setEnabled(False)
        layout.addWidget(self.scope_combo)

        format_label = QLabel("Format:")
        format_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(format_label)

        self.format_combo = QComboBox()
        self.format_combo.addItem("CSV", "csv")
        self.format_combo.addItem("XLSX", "xlsx")
        layout.addWidget(self.format_combo)

        ordering_checkbox = QCheckBox("Use current sort order")
        ordering_checkbox.setChecked(True)
        layout.addWidget(ordering_checkbox)
        self.ordering_checkbox = ordering_checkbox

        fields_label = QLabel("Fields to include:")
        fields_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(fields_label)

        self.field_checks: list[QCheckBox] = []
        field_grid = QGridLayout()
        layout.addLayout(field_grid)

        field_options = [
            ("ID", "id", True),
            ("License Plate", "license_plate", True),
            ("VIN", "vin", True),
            ("Vehicle (Year Make Model)", "vehicle", True),
            ("Cap", "capacity", True),
            ("Type", "type", True),
            ("Status", "status", True),
            ("Tags", "tags", True),
            ("Created", "created", True),
            ("Updated", "updated", True),
            ("Year", "year", False),
            ("Make", "make", False),
            ("Model", "model", False),
        ]

        for row, (label_text, field_key, checked) in enumerate(field_options):
            check = QCheckBox(label_text)
            check.setChecked(checked)
            check.setProperty("fieldKey", field_key)
            field_grid.addWidget(check, row // 2, row % 2)
            self.field_checks.append(check)

        layout.addStretch(1)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_scope(self) -> str:
        return str(self.scope_combo.currentData())

    def selected_format(self) -> str:
        return str(self.format_combo.currentData())

    def selected_fields(self) -> list[str]:
        return [
            str(check.property("fieldKey"))
            for check in self.field_checks
            if check.isChecked()
        ]

    def use_current_order(self) -> bool:
        return self.ordering_checkbox.isChecked()


# ---------------------------------------------------------------------------
# Import wizard
# ---------------------------------------------------------------------------


@dataclass
class ImportContext:
    repository: VehicleRepository
    type_lookup: dict[str, str]
    status_lookup: dict[str, str]
    file_path: Path | None = None
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)
    mapping: dict[str, str] = field(default_factory=dict)
    prepared_rows: list[dict[str, Any]] = field(default_factory=list)
    created_records: list[dict[str, Any]] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)


TARGET_FIELDS: list[tuple[str, str, bool]] = [
    ("id", "Vehicle ID", False),
    ("license_plate", "License Plate", True),
    ("vin", "VIN", True),
    ("year", "Year", False),
    ("make", "Make", False),
    ("model", "Model", False),
    ("capacity", "Capacity", False),
    ("type_id", "Type", False),
    ("status_id", "Status", False),
    ("tags", "Tags", False),
]


class UploadPage(QWizardPage):
    def __init__(self, context: ImportContext, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.context = context
        self.setTitle("Upload CSV/XLSX")
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        info = QLabel("Choose a CSV or Excel file to import. Headers are required for mapping.")
        info.setWordWrap(True)
        layout.addWidget(info)

        selector_layout = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setReadOnly(True)
        selector_layout.addWidget(self.path_edit)
        browse = QPushButton("Browse…")
        browse.clicked.connect(self._choose_file)
        selector_layout.addWidget(browse)
        layout.addLayout(selector_layout)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: #b71c1c;")
        layout.addWidget(self.error_label)

    def isComplete(self) -> bool:  # type: ignore[override]
        return bool(self.context.rows)

    def _choose_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select import file",
            "",
            "Data files (*.csv *.tsv *.xlsx);;All files (*.*)",
        )
        if not file_path:
            return
        path = Path(file_path)
        try:
            headers, rows = self._load_file(path)
        except Exception as exc:  # pragma: no cover - UI feedback
            self.error_label.setText(str(exc))
            self.context.rows.clear()
            self.context.headers.clear()
            self.completeChanged.emit()
            return

        self.context.file_path = path
        self.context.headers = headers
        self.context.rows = rows
        self.context.mapping.clear()
        self.context.prepared_rows.clear()
        self.context.errors.clear()
        self.context.created_records.clear()

        self.path_edit.setText(str(path))
        self.error_label.clear()
        self.completeChanged.emit()

    def _load_file(self, path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        ext = path.suffix.lower()
        if ext in {".csv", ".tsv"}:
            delimiter = "," if ext == ".csv" else "\t"
            with path.open("r", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh, delimiter=delimiter)
                headers = reader.fieldnames or []
                rows = [dict(row) for row in reader]
        elif ext == ".xlsx":
            from openpyxl import load_workbook  # type: ignore

            workbook = load_workbook(path, read_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
            except StopIteration:
                raise ValueError("The workbook is empty.")
            rows = []
            for line in rows_iter:
                row_dict = {headers[i]: line[i] for i in range(len(headers))}
                rows.append(row_dict)
        else:
            raise ValueError("Unsupported file type. Choose CSV or XLSX.")

        headers = [header.strip() for header in headers if header]
        if not headers:
            raise ValueError("The file does not include column headers.")
        if not rows:
            raise ValueError("No data rows found in the file.")
        return headers, rows


class MappingPage(QWizardPage):
    def __init__(self, context: ImportContext, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.context = context
        self.setTitle("Map Columns")
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        info = QLabel("Map columns from the file to vehicle fields. Required fields are highlighted.")
        info.setWordWrap(True)
        layout.addWidget(info)

        self.form_widget = QWidget()
        self.form_layout = QFormLayout(self.form_widget)
        self.form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(self.form_widget)

        layout.addStretch(1)

        self._combos: dict[str, QComboBox] = {}

    def initializePage(self) -> None:  # type: ignore[override]
        for combo in self._combos.values():
            combo.deleteLater()
        self._combos.clear()

        for key, label, required in TARGET_FIELDS:
            combo = QComboBox()
            combo.addItem("(Ignore)", "")
            for header in self.context.headers:
                combo.addItem(header, header)

            target = label.lower()
            best_index = -1
            for idx in range(1, combo.count()):
                header_text = combo.itemText(idx)
                if header_text.lower() == target or header_text.lower() == key.replace("_", " "):
                    best_index = idx
                    break
            if best_index >= 0:
                combo.setCurrentIndex(best_index)

            if required:
                combo.setProperty("required", True)
            self.form_layout.addRow(f"{label}:", combo)
            self._combos[key] = combo

        for key, value in self.context.mapping.items():
            combo = self._combos.get(key)
            if combo:
                index = combo.findData(value)
                if index >= 0:
                    combo.setCurrentIndex(index)

        self._update_styles()

    def isComplete(self) -> bool:  # type: ignore[override]
        for combo in self._combos.values():
            if combo.property("required") and not combo.currentData():
                return False
        return True

    def validatePage(self) -> bool:  # type: ignore[override]
        mapping: dict[str, str] = {}
        for key, combo in self._combos.items():
            value = combo.currentData()
            if combo.property("required") and not value:
                QMessageBox.warning(self, "Missing field", "Required fields must be mapped before continuing.")
                return False
            if value:
                mapping[key] = str(value)
        self.context.mapping = mapping
        return True

    def _update_styles(self) -> None:
        for combo in self._combos.values():
            if combo.property("required") and not combo.currentData():
                combo.setStyleSheet("border: 1px solid #c62828;")
            else:
                combo.setStyleSheet("")


class PreviewPage(QWizardPage):
    def __init__(self, context: ImportContext, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.context = context
        self.setTitle("Preview")
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        self.summary_label = QLabel("Previewing first 20 rows.")
        layout.addWidget(self.summary_label)

        self.table = QTableWidget(0, len(TARGET_FIELDS))
        headers = [label for _, label, _ in TARGET_FIELDS]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        layout.addWidget(self.table)

    def initializePage(self) -> None:  # type: ignore[override]
        mapping = self.context.mapping
        self.table.setRowCount(0)

        invalid_cells = 0
        preview_rows = self.context.rows[:20]
        for row_idx, source_row in enumerate(preview_rows):
            self.table.insertRow(row_idx)
            for col_idx, (key, _, required) in enumerate(TARGET_FIELDS):
                mapped = mapping.get(key)
                value = source_row.get(mapped, "") if mapped else ""
                display = "" if value is None else str(value)
                item = QTableWidgetItem(display)
                if required and not display.strip():
                    item.setBackground(QColor("#ffebee"))
                    invalid_cells += 1
                self.table.setItem(row_idx, col_idx, item)

        if invalid_cells:
            self.summary_label.setText(
                f"Previewing first 20 rows. {invalid_cells} required fields are blank and will prevent import."
            )
        else:
            self.summary_label.setText("Previewing first 20 rows.")

    def validatePage(self) -> bool:  # type: ignore[override]
        mapping = self.context.mapping
        prepared: list[dict[str, Any]] = []
        required_fields = {key for key, _, required in TARGET_FIELDS if required}

        for source_row in self.context.rows:
            payload: dict[str, Any] = {}
            invalid = False
            for key, _, required in TARGET_FIELDS:
                column = mapping.get(key)
                value = source_row.get(column) if column else None
                if isinstance(value, str):
                    cleaned: Any = value.strip()
                else:
                    cleaned = value

                if required and (cleaned in (None, "")):
                    invalid = True
                    break

                if cleaned in (None, ""):
                    continue

                if key == "capacity":
                    try:
                        payload[key] = int(cleaned)
                    except (ValueError, TypeError):
                        payload[key] = 0
                elif key == "year":
                    try:
                        payload[key] = int(cleaned)
                    except (ValueError, TypeError):
                        payload[key] = cleaned
                elif key == "tags":
                    payload[key] = [part.strip() for part in str(cleaned).replace(";", ",").split(",") if part.strip()]
                elif key in {"type_id", "status_id"}:
                    lookup = self.context.type_lookup if key == "type_id" else self.context.status_lookup
                    payload[key] = lookup.get(str(cleaned).lower(), str(cleaned))
                else:
                    payload[key] = cleaned

            if invalid:
                continue

            prepared.append(payload)

        if not prepared:
            QMessageBox.warning(self, "No valid rows", "No rows met the minimum requirements for import.")
            return False

        self.context.prepared_rows = prepared
        return True


class RunImportPage(QWizardPage):
    def __init__(self, context: ImportContext, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.context = context
        self.setTitle("Confirm & Import")
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        self.summary_label = QLabel("Ready to import.")
        layout.addWidget(self.summary_label)

        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        self.error_button = QPushButton("Download error CSV…")
        self.error_button.setEnabled(False)
        layout.addWidget(self.error_button)

        self.error_button.clicked.connect(self._download_errors)

    def initializePage(self) -> None:  # type: ignore[override]
        total = len(self.context.prepared_rows)
        self.summary_label.setText(f"Importing {total} rows…")
        self.progress.setValue(0)
        self.progress.setMaximum(total or 1)
        self.error_button.setEnabled(False)
        QApplication.processEvents()

        self.context.created_records.clear()
        self.context.errors.clear()

        for index, payload in enumerate(self.context.prepared_rows, start=1):
            try:
                response = self.context.repository.create_vehicle(payload)
                self.context.created_records.append(response)
            except Exception as exc:  # pragma: no cover - relies on database state
                error_entry = dict(payload)
                error_entry["error"] = str(exc)
                self.context.errors.append(error_entry)
            self.progress.setValue(index)
            QApplication.processEvents()

        if self.context.errors:
            self.summary_label.setText(
                f"Imported {len(self.context.created_records)} rows with {len(self.context.errors)} errors."
            )
            self.error_button.setEnabled(True)
        else:
            self.summary_label.setText(f"Successfully imported {len(self.context.created_records)} vehicles.")

        wizard = self.wizard()
        if wizard:
            wizard.button(QWizard.WizardButton.BackButton).setEnabled(False)
            wizard.button(QWizard.WizardButton.NextButton).setEnabled(False)
            wizard.button(QWizard.WizardButton.FinishButton).setEnabled(True)

    def isComplete(self) -> bool:  # type: ignore[override]
        return True

    def _download_errors(self) -> None:
        if not self.context.errors:
            return
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save error CSV",
            "vehicle-import-errors.csv",
            "CSV Files (*.csv)",
        )
        if not filename:
            return
        with open(filename, "w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(self.context.errors[0].keys()))
            writer.writeheader()
            for row in self.context.errors:
                writer.writerow(row)
        QMessageBox.information(self, "Saved", f"Error report saved to {filename}")


class VehicleImportWizard(QWizard):
    def __init__(
        self,
        *,
        repository: VehicleRepository,
        type_lookup: dict[str, str],
        status_lookup: dict[str, str],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Import Vehicles")
        self.setModal(True)
        self.setWizardStyle(QWizard.WizardStyle.ModernStyle)

        self.context = ImportContext(repository=repository, type_lookup=type_lookup, status_lookup=status_lookup)

        self._upload_page = UploadPage(self.context, self)
        self._mapping_page = MappingPage(self.context, self)
        self._preview_page = PreviewPage(self.context, self)
        self._run_page = RunImportPage(self.context, self)

        self.addPage(self._upload_page)
        self.addPage(self._mapping_page)
        self.addPage(self._preview_page)
        self.addPage(self._run_page)

        self.button(QWizard.WizardButton.FinishButton).setEnabled(False)

    def created_records(self) -> list[dict[str, Any]]:
        return list(self.context.created_records)


class VehicleInventoryPanel(QWidget):
    """Primary widget that renders the vehicle inventory management UI."""

    def __init__(self, parent: QWidget | None = None, repository: VehicleRepository | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("vehicleInventoryPanel")
        self.repository = repository or VehicleRepository()
        self._notifier = get_notifier()

        self._type_labels: dict[str, str] = {}
        self._status_labels: dict[str, str] = {}
        self._type_reverse: dict[str, str] = {}
        self._status_reverse: dict[str, str] = {}

        self._page = 1
        self._page_size = self._load_page_size()
        self._total_records = 0
        self._sort_key = "id"
        self._sort_order = "asc"
        self._selected_vehicle_id: str | None = None
        self._export_watcher: QFutureWatcher | None = None
        self._export_worker: _ExportWorkerThread | None = None

        self._setup_ui()
        self._load_reference_data()
        self.refresh(reset_page=True)

    # ----- UI construction -------------------------------------------------
    def _setup_ui(self) -> None:
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)

        card = QFrame(self)
        card.setObjectName("vehicleInventoryCard")
        card.setStyleSheet(
            """
            #vehicleInventoryCard {
                border-radius: 16px;
                background: palette(Base);
                border: 1px solid palette(Midlight);
            }
            QTableView {
                border: none;
                selection-background-color: palette(Highlight);
            }
            """
        )
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 20, 20, 20)
        card_layout.setSpacing(16)

        header_layout = QHBoxLayout()
        title = QLabel("Vehicle Inventory Management")
        title.setStyleSheet("font-size: 22px; font-weight: 700;")
        header_layout.addWidget(title)
        header_layout.addStretch(1)

        self.add_button = QPushButton("Add")
        self.add_button.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogNewFolder))
        self.add_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.add_button.setToolTip("Add a vehicle")
        header_layout.addWidget(self.add_button)

        self.import_button = QPushButton("Import")
        self.import_button.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowDown))
        self.import_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.import_button.setToolTip("Import vehicles from CSV/XLSX")
        header_layout.addWidget(self.import_button)

        self.export_button = QPushButton("Export")
        self.export_button.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.export_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.export_button.setToolTip("Export vehicles")
        header_layout.addWidget(self.export_button)

        card_layout.addLayout(header_layout)

        self.filter_bar = VehicleFilterBar(self)
        card_layout.addWidget(self.filter_bar)

        self.error_banner = QFrame()
        self.error_banner.setObjectName("vehicleErrorBanner")
        self.error_banner.setStyleSheet(
            """
            #vehicleErrorBanner {
                background: #fdecea;
                border-radius: 8px;
                border: 1px solid #f5c6cb;
            }
            """
        )
        self.error_banner.hide()
        error_layout = QHBoxLayout(self.error_banner)
        error_layout.setContentsMargins(12, 8, 12, 8)
        error_layout.setSpacing(12)
        self.error_label = QLabel("Failed to load data.")
        self.error_label.setStyleSheet("color: #b71c1c;")
        error_layout.addWidget(self.error_label, stretch=1)
        self.retry_button = QPushButton("Retry")
        error_layout.addWidget(self.retry_button)
        card_layout.addWidget(self.error_banner)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.splitter.setChildrenCollapsible(False)
        card_layout.addWidget(self.splitter, stretch=1)

        # Left pane with table and pagination
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        self.model = VehicleInventoryModel(self)
        self.table_view = QTableView()
        self.table_view.setObjectName("vehicleTable")
        self.table_view.setModel(self.model)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_view.setSortingEnabled(False)
        self.table_view.verticalHeader().setVisible(False)
        self.table_view.verticalHeader().setDefaultSectionSize(44)
        header = self.table_view.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionsClickable(True)
        header.setSortIndicatorShown(True)
        header.setSortIndicator(0, Qt.SortOrder.AscendingOrder)

        self.table_view.setItemDelegateForColumn(6, StatusPillDelegate(self.table_view))
        self.table_view.setItemDelegateForColumn(7, TagsChipDelegate(self.table_view))

        self.table_stack = QStackedLayout()
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(self.table_view)
        self.table_stack.addWidget(table_container)

        # Empty state when filters active
        self.no_results_widget = QWidget()
        nr_layout = QVBoxLayout(self.no_results_widget)
        nr_layout.addStretch(1)
        nr_message = QLabel("No vehicles match your filters.")
        nr_message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        nr_message.setStyleSheet("font-size: 16px; color: palette(Mid);")
        nr_layout.addWidget(nr_message)
        clear_filters_btn = QPushButton("Clear filters")
        clear_filters_btn.setFixedWidth(160)
        clear_filters_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_filters_btn.clicked.connect(self.filter_bar.reset_filters)
        clear_filters_btn.setStyleSheet("padding: 6px 12px;")
        nr_layout.addWidget(clear_filters_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        nr_layout.addStretch(1)
        self.table_stack.addWidget(self.no_results_widget)

        # First run empty state
        self.first_run_widget = QWidget()
        fr_layout = QVBoxLayout(self.first_run_widget)
        fr_layout.addStretch(1)
        fr_message = QLabel("Add your first vehicle.")
        fr_message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        fr_message.setStyleSheet("font-size: 16px; color: palette(Mid);")
        fr_layout.addWidget(fr_message)
        fr_add_button = QPushButton("Add vehicle")
        fr_add_button.setFixedWidth(160)
        fr_add_button.clicked.connect(self._on_add_clicked)
        fr_add_button.setCursor(Qt.CursorShape.PointingHandCursor)
        fr_layout.addWidget(fr_add_button, alignment=Qt.AlignmentFlag.AlignCenter)
        fr_layout.addStretch(1)
        self.table_stack.addWidget(self.first_run_widget)

        left_layout.addLayout(self.table_stack, stretch=1)

        self.pagination = PaginationControls(self)
        left_layout.addWidget(self.pagination)

        self.splitter.addWidget(left_widget)

        # Detail panel on the right
        self.detail_panel = VehicleDetailPanel(self)
        self.splitter.addWidget(self.detail_panel)
        self.splitter.setStretchFactor(0, 3)
        self.splitter.setStretchFactor(1, 2)

        main_layout.addWidget(card)

        # Signals
        self.filter_bar.filtersChanged.connect(lambda: self.refresh(reset_page=True))
        self.pagination.pageRequested.connect(self._on_page_requested)
        self.pagination.pageSizeChanged.connect(self._on_page_size_changed)
        self.add_button.clicked.connect(self._on_add_clicked)
        self.import_button.clicked.connect(self._on_import_clicked)
        self.export_button.clicked.connect(self._on_export_clicked)
        self.retry_button.clicked.connect(self._on_retry)
        header.sectionClicked.connect(self._on_header_clicked)
        self.table_view.activated.connect(self._open_selected_for_edit)

        selection_model = self.table_view.selectionModel()
        if selection_model:
            selection_model.currentChanged.connect(self._on_current_changed)

        self.search_shortcut = QShortcut(QKeySequence.StandardKey.Find, self)
        self.search_shortcut.activated.connect(self.filter_bar.focus_search)

    # ----- Persistence helpers -------------------------------------------
    def _load_page_size(self) -> int:
        settings = QSettings()
        value = settings.value("vehicle_inventory/page_size", 20)
        try:
            return max(5, int(value))
        except (TypeError, ValueError):
            return 20

    def _save_page_size(self, value: int) -> None:
        settings = QSettings()
        settings.setValue("vehicle_inventory/page_size", int(value))

    # ----- Data loading ---------------------------------------------------
    def _load_reference_data(self) -> None:
        try:
            types = self.repository.list_vehicle_types()
            statuses = self.repository.list_statuses()
        except Exception as exc:  # pragma: no cover - relies on runtime DB
            self._show_error(f"Unable to load reference data: {exc}")
            return

        self._hide_error()
        self.filter_bar.apply_reference_options(types, statuses)

        self._type_labels = {str(entry.get("id")): str(entry.get("name")) for entry in types}
        self._status_labels = {str(entry.get("id")): str(entry.get("name")) for entry in statuses}
        self._type_reverse = {str(entry.get("name")).lower(): str(entry.get("id")) for entry in types}
        self._type_reverse.update({str(entry.get("id")).lower(): str(entry.get("id")) for entry in types})
        self._status_reverse = {str(entry.get("name")).lower(): str(entry.get("id")) for entry in statuses}
        self._status_reverse.update({str(entry.get("id")).lower(): str(entry.get("id")) for entry in statuses})

    # ----- Refresh logic --------------------------------------------------
    def refresh(self, reset_page: bool = False) -> None:
        if reset_page:
            self._page = 1

        filters = self.filter_bar.filters()
        try:
            rows, total = self.repository.list_inventory(
                search=filters.get("search"),
                type_filter=filters.get("type"),
                status_filter=filters.get("status"),
                sort_key=self._sort_key,
                sort_order=self._sort_order,
                offset=(self._page - 1) * self._page_size,
                limit=self._page_size,
            )
        except Exception as exc:  # pragma: no cover - runtime DB dependent
            self._show_error(f"Unable to load vehicles: {exc}")
            return

        self._hide_error()
        self._total_records = total

        max_page = max(1, math.ceil(total / self._page_size)) if self._page_size else 1
        if self._page > max_page:
            self._page = max_page
            if total > 0:
                self.refresh(reset_page=False)
                return

        records = [self._build_record(row) for row in rows]
        self.model.set_records(records)
        self.pagination.update_state(total=total, page=self._page, page_size=self._page_size)

        if total == 0:
            if filters.get("search") or filters.get("type") or filters.get("status"):
                self.table_stack.setCurrentWidget(self.no_results_widget)
            else:
                self.table_stack.setCurrentWidget(self.first_run_widget)
            self.detail_panel.set_record(None)
        else:
            self.table_stack.setCurrentIndex(0)
            self._restore_selection()

    def _build_record(self, payload: dict[str, Any]) -> VehicleRecord:
        return self._create_vehicle_record(payload, self._type_labels, self._status_labels)

    @staticmethod
    def _create_vehicle_record(
        payload: dict[str, Any],
        type_labels: dict[str, str],
        status_labels: dict[str, str],
    ) -> VehicleRecord:
        identifier = str(coalesce([payload.get("id"), payload.get("identifier")]) or "")
        license_plate = payload.get("license_plate")
        vin = payload.get("vin")
        year_value = payload.get("year")
        try:
            year = int(year_value) if year_value not in (None, "") else None
        except (ValueError, TypeError):
            year = None
        make = payload.get("make")
        model = payload.get("model")
        capacity_value = payload.get("capacity")
        try:
            capacity = int(capacity_value) if capacity_value not in (None, "") else None
        except (ValueError, TypeError):
            capacity = None

        type_id_raw = coalesce([payload.get("type_id"), payload.get("type")])
        type_id = str(type_id_raw).strip() if type_id_raw not in (None, "") else None
        status_id_raw = coalesce([payload.get("status_id"), payload.get("status")])
        status_id = str(status_id_raw).strip() if status_id_raw not in (None, "") else None

        type_label = type_labels.get(type_id) if type_id else None
        status_label = status_labels.get(status_id) if status_id else None

        created_value = coalesce([payload.get("created_at"), payload.get("created_ts"), payload.get("created")])
        updated_value = coalesce([payload.get("updated_at"), payload.get("updated_ts"), payload.get("updated")])

        return VehicleRecord(
            identifier=identifier,
            license_plate=str(license_plate) if license_plate not in (None, "") else None,
            vin=str(vin) if vin not in (None, "") else None,
            year=year,
            make=str(make) if make not in (None, "") else None,
            model=str(model) if model not in (None, "") else None,
            capacity=capacity,
            type_id=type_id,
            type_label=type_label,
            status_id=status_id,
            status_label=status_label,
            tags=ensure_list(payload.get("tags")),
            created=format_timestamp(created_value),
            updated=format_timestamp(updated_value),
            organization=payload.get("organization"),
            raw=dict(payload),
        )

    # ----- Selection handling --------------------------------------------
    def _restore_selection(self) -> None:
        selection_model = self.table_view.selectionModel()
        if selection_model is None:
            return
        if self._selected_vehicle_id:
            for row in range(self.model.rowCount()):
                record = self.model.record(row)
                if record and str(record.identifier) == str(self._selected_vehicle_id):
                    index = self.model.index(row, 0)
                    selection_model.select(
                        index,
                        QItemSelectionModel.SelectionFlag.ClearAndSelect | QItemSelectionModel.SelectionFlag.Rows,
                    )
                    self.table_view.scrollTo(index, QTableView.ScrollHint.PositionAtCenter)
                    self.detail_panel.set_record(record)
                    return
        if self.model.rowCount() > 0:
            index = self.model.index(0, 0)
            selection_model.select(
                index,
                QItemSelectionModel.SelectionFlag.ClearAndSelect | QItemSelectionModel.SelectionFlag.Rows,
            )
            self.table_view.scrollTo(index, QTableView.ScrollHint.PositionAtCenter)
            self.detail_panel.set_record(self.model.record(0))
        else:
            self.detail_panel.set_record(None)

    def _on_current_changed(self, current: QModelIndex, _: QModelIndex) -> None:
        record = current.data(RECORD_ROLE)
        if isinstance(record, VehicleRecord):
            self._selected_vehicle_id = str(record.identifier)
            self.detail_panel.set_record(record)
        else:
            self.detail_panel.set_record(None)

    # ----- Pagination & sorting ------------------------------------------
    def _on_page_requested(self, page: int) -> None:
        max_page = max(1, math.ceil(self._total_records / self._page_size)) if self._page_size else 1
        page = min(max(1, page), max_page)
        if page == self._page:
            return
        self._page = page
        self.refresh(reset_page=False)

    def _on_page_size_changed(self, page_size: int) -> None:
        if page_size <= 0:
            return
        self._page_size = page_size
        self._save_page_size(page_size)
        self._page = 1
        self.refresh(reset_page=False)

    def _on_header_clicked(self, section: int) -> None:
        sort_key = self.model.sort_key_for_column(section)
        if sort_key == self._sort_key:
            self._sort_order = "desc" if self._sort_order == "asc" else "asc"
        else:
            self._sort_key = sort_key
            self._sort_order = "asc"
        order = Qt.SortOrder.AscendingOrder if self._sort_order == "asc" else Qt.SortOrder.DescendingOrder
        self.table_view.horizontalHeader().setSortIndicator(section, order)
        self.refresh(reset_page=False)

    # ----- Error banner ---------------------------------------------------
    def _show_error(self, message: str) -> None:
        self.error_label.setText(message)
        self.error_banner.show()

    def _hide_error(self) -> None:
        self.error_banner.hide()

    def _on_retry(self) -> None:
        if not self._type_labels or not self._status_labels:
            self._load_reference_data()
        self.refresh(reset_page=False)

    # ----- Actions --------------------------------------------------------
    def _open_selected_for_edit(self, index: QModelIndex) -> None:
        record = index.data(RECORD_ROLE)
        if isinstance(record, VehicleRecord):
            raw_id = record.raw.get("id")
            try:
                vehicle_id = int(raw_id)
            except (TypeError, ValueError):
                vehicle_id = None
            dialog = VehicleEditDialog(vehicle_id=vehicle_id, repository=self.repository, parent=self)
            dialog.vehicleSaved.connect(self._on_vehicle_saved)
            dialog.exec()

    def _on_add_clicked(self) -> None:
        dialog = VehicleEditDialog(repository=self.repository, parent=self)
        dialog.vehicleSaved.connect(self._on_vehicle_saved)
        dialog.exec()

    def _on_vehicle_saved(self, record: dict[str, Any]) -> None:
        vehicle_id = record.get("id")
        if vehicle_id is None:
            self.refresh(reset_page=False)
            return
        self._selected_vehicle_id = str(vehicle_id)
        self._locate_and_select_vehicle(str(vehicle_id))
        self._show_toast("Vehicle saved", f"Vehicle #{vehicle_id} saved successfully.")

    def _locate_and_select_vehicle(self, vehicle_id: str) -> None:
        filters = self.filter_bar.filters()
        try:
            rows, _ = self.repository.list_inventory(
                search=filters.get("search"),
                type_filter=filters.get("type"),
                status_filter=filters.get("status"),
                sort_key=self._sort_key,
                sort_order=self._sort_order,
                offset=0,
                limit=None,
            )
        except Exception:
            self.refresh(reset_page=False)
            return

        target_index = 0
        for idx, row in enumerate(rows):
            if str(row.get("id")) == str(vehicle_id):
                target_index = idx
                break
        self._page = (target_index // self._page_size) + 1 if self._page_size else 1
        self.refresh(reset_page=False)

    def _on_import_clicked(self) -> None:
        wizard = VehicleImportWizard(
            repository=self.repository,
            type_lookup=self._type_reverse,
            status_lookup=self._status_reverse,
            parent=self,
        )
        result = wizard.exec()
        created = wizard.created_records()
        errors = len(wizard.context.errors)
        if result == QDialog.DialogCode.Accepted and (created or errors):
            if created:
                self._selected_vehicle_id = str(created[-1].get("id")) if created[-1].get("id") is not None else None
            self.refresh(reset_page=True)
            if created and errors:
                self._show_toast("Import complete", f"{len(created)} vehicles imported with {errors} errors.", severity="warning")
            elif created:
                self._show_toast("Import complete", f"Imported {len(created)} vehicles.")
            elif errors:
                self._show_toast("Import issues", "No rows were imported. Check the error report.", severity="warning")

    def _on_export_clicked(self) -> None:
        selection_model = self.table_view.selectionModel()
        allow_selected = bool(selection_model and selection_model.hasSelection())
        dialog = VehicleExportDialog(parent=self, allow_selected=allow_selected)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        scope = dialog.selected_scope()
        file_format = dialog.selected_format()
        fields = dialog.selected_fields()
        use_order = dialog.use_current_order()

        selected_records: list[dict[str, Any]] = []
        if scope == "selected":
            if not selection_model or not selection_model.hasSelection():
                QMessageBox.information(self, "No selection", "Select rows to export or choose a different scope.")
                return
            rows = sorted({index.row() for index in selection_model.selectedRows()})
            for row in rows:
                record = self.model.record(row)
                if record:
                    selected_records.append(record.raw)

        params = {
            "db_path": getattr(self.repository, "_db_path", None),
            "filters": self.filter_bar.filters(),
            "sort_key": self._sort_key if use_order else "id",
            "sort_order": self._sort_order if use_order else "asc",
            "scope": scope,
            "fields": fields,
            "format": file_format,
            "selected": selected_records,
            "type_labels": self._type_labels,
            "status_labels": self._status_labels,
        }

        self.export_button.setEnabled(False)

        run_fn = getattr(QtConcurrent, "run", None)
        if QFutureWatcher is not None and callable(run_fn):
            watcher = QFutureWatcher(self)
            future = run_fn(self._perform_export, params)
            watcher.setFuture(future)
            watcher.finished.connect(lambda: self._on_export_finished(watcher))
            self._export_watcher = watcher
        else:
            worker = _ExportWorkerThread(self._perform_export, params, self)
            worker.completed.connect(self._on_export_finished)
            worker.failed.connect(self._on_export_failed)
            worker.finished.connect(lambda: self._on_export_worker_finished(worker))
            worker.start()
            self._export_worker = worker

    @staticmethod
    def _perform_export(params: dict[str, Any]) -> dict[str, Any]:
        scope = params.get("scope", "all")
        fmt = params.get("format", "csv")
        fields: list[str] = params.get("fields", [])
        type_labels = params.get("type_labels", {})
        status_labels = params.get("status_labels", {})

        repository = VehicleRepository(params.get("db_path"))

        if scope == "selected":
            rows = params.get("selected", [])
        else:
            search = params.get("filters", {}).get("search") if scope == "filters" else None
            type_filter = params.get("filters", {}).get("type") if scope == "filters" else None
            status_filter = params.get("filters", {}).get("status") if scope == "filters" else None
            rows, _ = repository.list_inventory(
                search=search,
                type_filter=type_filter,
                status_filter=status_filter,
                sort_key=params.get("sort_key", "id"),
                sort_order=params.get("sort_order", "asc"),
                offset=0,
                limit=None,
            )

        records = [
            VehicleInventoryPanel._create_vehicle_record(row, type_labels, status_labels) for row in rows
        ]

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        scope_slug = scope.replace(" ", "-")
        filename = f"vehicles-{scope_slug}-{timestamp}.{fmt}"
        path = Path(tempfile.gettempdir()) / filename

        VehicleInventoryPanel._write_export_file(path, records, fields, fmt)

        return {"path": str(path), "count": len(records), "scope": scope, "format": fmt}

    @staticmethod
    def _write_export_file(path: Path, records: Sequence[VehicleRecord], fields: Sequence[str], fmt: str) -> None:
        headers = [VehicleInventoryPanel._field_label(field) for field in fields]
        if fmt == "xlsx":
            from openpyxl import Workbook  # type: ignore

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Vehicles"
            sheet.append(headers)
            for record in records:
                sheet.append(VehicleInventoryPanel._render_row(record, fields))
            workbook.save(path)
        else:
            with path.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(headers)
                for record in records:
                    writer.writerow(VehicleInventoryPanel._render_row(record, fields))

    @staticmethod
    def _field_label(field: str) -> str:
        mapping = {
            "id": "ID",
            "license_plate": "License Plate",
            "vin": "VIN",
            "vehicle": "Vehicle (Year Make Model)",
            "capacity": "Cap",
            "type": "Type",
            "status": "Status",
            "tags": "Tags",
            "created": "Created",
            "updated": "Updated",
            "year": "Year",
            "make": "Make",
            "model": "Model",
        }
        return mapping.get(field, field.title())

    @staticmethod
    def _render_row(record: VehicleRecord, fields: Sequence[str]) -> list[str]:
        values: list[str] = []
        for field in fields:
            if field == "id":
                values.append(str(record.identifier))
            elif field == "license_plate":
                values.append(record.license_plate or "")
            elif field == "vin":
                values.append(record.vin or "")
            elif field == "vehicle":
                values.append(record.vehicle_label())
            elif field == "capacity":
                values.append(str(record.capacity) if record.capacity not in (None, "") else "")
            elif field == "type":
                values.append(record.type_label or "")
            elif field == "status":
                values.append(record.status_label or "")
            elif field == "tags":
                values.append(", ".join(record.tags))
            elif field == "created":
                values.append(record.created or "")
            elif field == "updated":
                values.append(record.updated or "")
            elif field == "year":
                values.append(str(record.year) if record.year not in (None, "") else "")
            elif field == "make":
                values.append(record.make or "")
            elif field == "model":
                values.append(record.model or "")
            else:
                values.append(record.raw.get(field, ""))
        return values

    def _on_export_finished(self, result_or_watcher: Any) -> None:
        self.export_button.setEnabled(True)

        if hasattr(result_or_watcher, "result"):
            if result_or_watcher is self._export_watcher:
                self._export_watcher = None
            try:
                result = result_or_watcher.result()
            except Exception as exc:  # pragma: no cover - runtime dependent
                self._show_toast("Export failed", str(exc), severity="error")
                return
        else:
            result = result_or_watcher

        self._present_export_result(result)

    def _on_export_failed(self, message: str) -> None:
        self.export_button.setEnabled(True)
        self._show_toast("Export failed", message, severity="error")

    def _on_export_worker_finished(self, worker: _ExportWorkerThread) -> None:
        if self._export_worker is worker:
            self._export_worker = None
        worker.deleteLater()

    def _present_export_result(self, result: dict[str, Any]) -> None:
        path = Path(result.get("path"))
        count = result.get("count", 0)
        scope = result.get("scope", "all")
        message = f"{count} vehicles exported ({scope}). Saved to {path}."
        self._show_toast("Export ready", message)
        QMessageBox.information(self, "Export ready", message)

    # ----- Permissions ----------------------------------------------------
    def set_actions_enabled(
        self,
        *,
        add: Optional[bool] = None,
        import_: Optional[bool] = None,
        export: Optional[bool] = None,
    ) -> None:
        if add is not None:
            self.add_button.setEnabled(add)
        if import_ is not None:
            self.import_button.setEnabled(import_)
        if export is not None:
            self.export_button.setEnabled(export)

    # ----- Toast helper ---------------------------------------------------
    def _show_toast(self, title: str, message: str, *, severity: str = "success") -> None:
        try:
            self._notifier.notify(
                Notification(
                    title=title,
                    message=message,
                    severity=severity if severity in {"info", "success", "warning", "error"} else "info",
                    source="Vehicle Inventory",
                )
            )
        except Exception:
            pass

    # ----- Qt overrides ---------------------------------------------------
    def resizeEvent(self, event: QEvent) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        if isinstance(event, QEvent):
            width = self.width()
            orientation = Qt.Orientation.Vertical if width < 900 else Qt.Orientation.Horizontal
            if self.splitter.orientation() != orientation:
                self.splitter.setOrientation(orientation)

