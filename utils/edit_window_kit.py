"""Shared building blocks for Edit-menu catalog windows.

Implements the reusable pieces of ``Design Documents/edit_window_style_guide.md``
(pill delegate, pagination footer, import wizard, export dialog, and an async
export runner) so individual Edit-menu windows configure field lists rather than
re-implementing this plumbing per module.

Background execution note: ``PySide6.QtConcurrent.run`` is not available in this
project's PySide6 build (confirmed: the attribute does not exist), and
``PySide6.QtCore.QFutureWatcher`` fails to import as well. ``run_async`` below
always uses a plain ``QThread`` worker — there is no QtConcurrent code path to
keep in sync with a fallback.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Optional, Sequence

from PySide6.QtCore import QModelIndex, QObject, Qt, QThread, QTimer, Signal
from PySide6.QtGui import QColor, QPainter, QPen
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QWizard,
    QWizardPage,
)

__all__ = [
    "FieldSpec",
    "PillDelegate",
    "PaginationControls",
    "ImportWizard",
    "ExportDialog",
    "run_async",
    "write_export_file",
]


@dataclass
class FieldSpec:
    key: str
    label: str
    required: bool = False
    in_export_default: bool = True


# ---------------------------------------------------------------------------
# Pill delegate — for status/condition-like columns
# ---------------------------------------------------------------------------


class PillDelegate(QStyledItemDelegate):
    """Render a column's display text as a rounded color pill."""

    def __init__(
        self,
        parent,
        color_map: dict[str, tuple[str, str]],
        default: tuple[str, str] = ("#5f6368", "#ffffff"),
    ) -> None:
        super().__init__(parent)
        self._color_map = {k.lower(): v for k, v in color_map.items()}
        self._default = default

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex) -> None:  # type: ignore[override]
        text = str(index.data(Qt.ItemDataRole.DisplayRole) or "")
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        bg, fg = self._color_map.get(text.lower(), self._default)
        rect = option.rect.adjusted(6, 8, -6, -8)
        radius = rect.height() / 2
        painter.setBrush(QColor(bg))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(rect, radius, radius)
        painter.setPen(QPen(QColor(fg)))
        font = option.font
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, text or "—")
        painter.restore()


# ---------------------------------------------------------------------------
# Pagination footer
# ---------------------------------------------------------------------------


class PaginationControls(QWidget):
    """Footer with status text, page-size selector, and prev/next navigation."""

    pageRequested = Signal(int)
    pageSizeChanged = Signal(int)

    def __init__(self, parent: QWidget | None = None, page_sizes: Sequence[int] = (20, 50, 100)) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.status_label = QLabel("0–0 of 0")
        layout.addWidget(self.status_label)
        layout.addStretch(1)

        self.page_size_combo = QComboBox()
        for size in page_sizes:
            self.page_size_combo.addItem(f"{size} / page", size)
        layout.addWidget(self.page_size_combo)

        self.prev_button = QToolButton()
        self.prev_button.setText("<")
        layout.addWidget(self.prev_button)

        self.next_button = QToolButton()
        self.next_button.setText(">")
        layout.addWidget(self.next_button)

        self.prev_button.clicked.connect(lambda: self.pageRequested.emit(self._current_page - 1))
        self.next_button.clicked.connect(lambda: self.pageRequested.emit(self._current_page + 1))
        self.page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)

        self._current_page = 1
        self._total_pages = 1

    def update_state(self, *, total: int, page: int, page_size: int) -> None:
        import math

        self._current_page = max(1, page)
        self.page_size_combo.blockSignals(True)
        idx = self.page_size_combo.findData(page_size)
        if idx >= 0:
            self.page_size_combo.setCurrentIndex(idx)
        self.page_size_combo.blockSignals(False)

        if total == 0:
            self.status_label.setText("0 of 0")
            self.prev_button.setEnabled(False)
            self.next_button.setEnabled(False)
            return

        total_pages = max(1, math.ceil(total / page_size)) if page_size else 1
        self._total_pages = total_pages
        page = min(max(1, page), total_pages)
        start = (page - 1) * page_size + 1
        end = min(total, page * page_size)
        self.status_label.setText(f"{start}–{end} of {total}")
        self.prev_button.setEnabled(page > 1)
        self.next_button.setEnabled(page < total_pages)

    def _on_page_size_changed(self) -> None:
        size = self.page_size_combo.currentData()
        if size:
            self.pageSizeChanged.emit(int(size))


# ---------------------------------------------------------------------------
# Import wizard: Upload -> Map columns -> Preview -> Run
# ---------------------------------------------------------------------------


class _ImportContext:
    def __init__(self, fields: Sequence[FieldSpec], import_row: Callable[[dict[str, Any]], Any]):
        self.fields = list(fields)
        self.import_row = import_row
        self.headers: list[str] = []
        self.rows: list[dict[str, Any]] = []
        self.mapping: dict[str, str] = {}
        self.prepared_rows: list[dict[str, Any]] = []
        self.created: list[Any] = []
        self.errors: list[dict[str, Any]] = []


class _UploadPage(QWizardPage):
    def __init__(self, ctx: _ImportContext, parent=None) -> None:
        super().__init__(parent)
        self.ctx = ctx
        self.setTitle("Upload CSV/XLSX")
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Choose a CSV or Excel file to import. Headers are required for mapping."))

        row = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setReadOnly(True)
        row.addWidget(self.path_edit)
        browse = QPushButton("Browse…")
        browse.clicked.connect(self._choose_file)
        row.addWidget(browse)
        layout.addLayout(row)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: #b71c1c;")
        layout.addWidget(self.error_label)

    def isComplete(self) -> bool:  # type: ignore[override]
        return bool(self.ctx.rows)

    def _choose_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select import file", "", "Data files (*.csv *.tsv *.xlsx);;All files (*.*)"
        )
        if not file_path:
            return
        path = Path(file_path)
        try:
            headers, rows = _load_table_file(path)
        except Exception as exc:
            self.error_label.setText(str(exc))
            self.ctx.rows.clear()
            self.ctx.headers.clear()
            self.completeChanged.emit()
            return

        self.ctx.headers = headers
        self.ctx.rows = rows
        self.ctx.mapping.clear()
        self.ctx.prepared_rows.clear()
        self.ctx.errors.clear()
        self.ctx.created.clear()

        self.path_edit.setText(str(path))
        self.error_label.clear()
        self.completeChanged.emit()


def _load_table_file(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv"}:
        delimiter = "," if ext == ".csv" else "\t"
        with path.open("r", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            headers = reader.fieldnames or []
            rows = [dict(row) for row in reader]
    elif ext == ".xlsx":
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
        except StopIteration:
            raise ValueError("The workbook is empty.")
        rows = []
        for line in rows_iter:
            rows.append({headers[i]: line[i] for i in range(len(headers))})
    else:
        raise ValueError("Unsupported file type. Choose CSV or XLSX.")

    headers = [h.strip() for h in headers if h]
    if not headers:
        raise ValueError("The file does not include column headers.")
    if not rows:
        raise ValueError("No data rows found in the file.")
    return headers, rows


class _MappingPage(QWizardPage):
    def __init__(self, ctx: _ImportContext, parent=None) -> None:
        super().__init__(parent)
        self.ctx = ctx
        self.setTitle("Map Columns")
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Map columns from the file to fields. Required fields are highlighted."))

        self.form_widget = QWidget()
        self.form_layout = QFormLayout(self.form_widget)
        self.form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(self.form_widget)
        layout.addStretch(1)

        self._combos: dict[str, QComboBox] = {}

    def initializePage(self) -> None:  # type: ignore[override]
        for combo in self._combos.values():
            combo.deleteLater()
        self._combos.clear()

        for spec in self.ctx.fields:
            combo = QComboBox()
            combo.addItem("(Ignore)", "")
            for header in self.ctx.headers:
                combo.addItem(header, header)

            best_index = -1
            target = spec.label.lower()
            for idx in range(1, combo.count()):
                text = combo.itemText(idx).lower()
                if text == target or text == spec.key.replace("_", " "):
                    best_index = idx
                    break
            if best_index >= 0:
                combo.setCurrentIndex(best_index)

            combo.setProperty("required", spec.required)
            self.form_layout.addRow(f"{spec.label}:", combo)
            self._combos[spec.key] = combo

        self._update_styles()
        for combo in self._combos.values():
            combo.currentIndexChanged.connect(self._update_styles)
            combo.currentIndexChanged.connect(self.completeChanged)

    def isComplete(self) -> bool:  # type: ignore[override]
        for combo in self._combos.values():
            if combo.property("required") and not combo.currentData():
                return False
        return True

    def validatePage(self) -> bool:  # type: ignore[override]
        mapping: dict[str, str] = {}
        for key, combo in self._combos.items():
            value = combo.currentData()
            if combo.property("required") and not value:
                QMessageBox.warning(self, "Missing field", "Required fields must be mapped before continuing.")
                return False
            if value:
                mapping[key] = str(value)
        self.ctx.mapping = mapping
        return True

    def _update_styles(self) -> None:
        for combo in self._combos.values():
            if combo.property("required") and not combo.currentData():
                combo.setStyleSheet("border: 1px solid #c62828;")
            else:
                combo.setStyleSheet("")


class _PreviewPage(QWizardPage):
    def __init__(self, ctx: _ImportContext, parent=None) -> None:
        super().__init__(parent)
        self.ctx = ctx
        self.setTitle("Preview")
        layout = QVBoxLayout(self)
        self.summary_label = QLabel("Previewing first 20 rows.")
        layout.addWidget(self.summary_label)

        self.table = QTableWidget(0, len(ctx.fields))
        self.table.setHorizontalHeaderLabels([spec.label for spec in ctx.fields])
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        layout.addWidget(self.table)

    def initializePage(self) -> None:  # type: ignore[override]
        mapping = self.ctx.mapping
        self.table.setRowCount(0)
        invalid_cells = 0
        preview_rows = self.ctx.rows[:20]
        for row_idx, source_row in enumerate(preview_rows):
            self.table.insertRow(row_idx)
            for col_idx, spec in enumerate(self.ctx.fields):
                mapped = mapping.get(spec.key)
                value = source_row.get(mapped, "") if mapped else ""
                display = "" if value is None else str(value)
                item = QTableWidgetItem(display)
                if spec.required and not display.strip():
                    item.setBackground(QColor("#ffebee"))
                    invalid_cells += 1
                self.table.setItem(row_idx, col_idx, item)

        if invalid_cells:
            self.summary_label.setText(
                f"Previewing first 20 rows. {invalid_cells} required fields are blank and will prevent import."
            )
        else:
            self.summary_label.setText("Previewing first 20 rows.")

    def validatePage(self) -> bool:  # type: ignore[override]
        mapping = self.ctx.mapping
        prepared: list[dict[str, Any]] = []
        for source_row in self.ctx.rows:
            payload: dict[str, Any] = {}
            invalid = False
            for spec in self.ctx.fields:
                column = mapping.get(spec.key)
                value = source_row.get(column) if column else None
                cleaned = value.strip() if isinstance(value, str) else value
                if spec.required and cleaned in (None, ""):
                    invalid = True
                    break
                if cleaned not in (None, ""):
                    payload[spec.key] = cleaned
            if not invalid:
                prepared.append(payload)

        if not prepared:
            QMessageBox.warning(self, "No valid rows", "No rows met the minimum requirements for import.")
            return False

        self.ctx.prepared_rows = prepared
        return True


class _RunImportPage(QWizardPage):
    def __init__(self, ctx: _ImportContext, parent=None) -> None:
        super().__init__(parent)
        self.ctx = ctx
        self.setTitle("Confirm & Import")
        layout = QVBoxLayout(self)
        self.summary_label = QLabel("Ready to import.")
        layout.addWidget(self.summary_label)

        from PySide6.QtWidgets import QProgressBar

        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        self.error_button = QPushButton("Download error CSV…")
        self.error_button.setEnabled(False)
        self.error_button.clicked.connect(self._download_errors)
        layout.addWidget(self.error_button)

    def initializePage(self) -> None:  # type: ignore[override]
        from PySide6.QtWidgets import QApplication

        total = len(self.ctx.prepared_rows)
        self.summary_label.setText(f"Importing {total} rows…")
        self.progress.setValue(0)
        self.progress.setMaximum(total or 1)
        self.error_button.setEnabled(False)
        QApplication.processEvents()

        self.ctx.created.clear()
        self.ctx.errors.clear()

        for index, payload in enumerate(self.ctx.prepared_rows, start=1):
            try:
                result = self.ctx.import_row(payload)
                self.ctx.created.append(result if result is not None else payload)
            except Exception as exc:
                error_entry = dict(payload)
                error_entry["error"] = str(exc)
                self.ctx.errors.append(error_entry)
            self.progress.setValue(index)
            QApplication.processEvents()

        if self.ctx.errors:
            self.summary_label.setText(
                f"Imported {len(self.ctx.created)} rows with {len(self.ctx.errors)} errors."
            )
            self.error_button.setEnabled(True)
        else:
            self.summary_label.setText(f"Successfully imported {len(self.ctx.created)} rows.")

        wizard = self.wizard()
        if wizard:
            wizard.button(QWizard.WizardButton.BackButton).setEnabled(False)
            wizard.button(QWizard.WizardButton.NextButton).setEnabled(False)
            wizard.button(QWizard.WizardButton.FinishButton).setEnabled(True)

    def isComplete(self) -> bool:  # type: ignore[override]
        return True

    def _download_errors(self) -> None:
        if not self.ctx.errors:
            return
        filename, _ = QFileDialog.getSaveFileName(self, "Save error CSV", "import-errors.csv", "CSV Files (*.csv)")
        if not filename:
            return
        with open(filename, "w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(self.ctx.errors[0].keys()))
            writer.writeheader()
            for row in self.ctx.errors:
                writer.writerow(row)
        QMessageBox.information(self, "Saved", f"Error report saved to {filename}")


class ImportWizard(QWizard):
    """Generic Upload -> Map -> Preview -> Run import wizard.

    ``import_row`` is called once per prepared row with the mapped payload dict;
    it should perform the actual create call (e.g. an API POST) and may raise to
    record a per-row error.
    """

    def __init__(
        self,
        *,
        fields: Sequence[FieldSpec],
        import_row: Callable[[dict[str, Any]], Any],
        title: str = "Import",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setWizardStyle(QWizard.WizardStyle.ModernStyle)

        self.ctx = _ImportContext(fields, import_row)
        self.addPage(_UploadPage(self.ctx, self))
        self.addPage(_MappingPage(self.ctx, self))
        self.addPage(_PreviewPage(self.ctx, self))
        self.addPage(_RunImportPage(self.ctx, self))
        self.button(QWizard.WizardButton.FinishButton).setEnabled(False)

    def created_records(self) -> list[Any]:
        return list(self.ctx.created)

    def error_count(self) -> int:
        return len(self.ctx.errors)


# ---------------------------------------------------------------------------
# Export dialog: scope / format / fields / sort order
# ---------------------------------------------------------------------------


class ExportDialog(QDialog):
    """Scope + format + field-selection dialog used before running an export."""

    def __init__(
        self,
        *,
        fields: Sequence[FieldSpec],
        allow_selected: bool = True,
        title: str = "Export",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(420)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        scope_label = QLabel("Scope:")
        scope_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(scope_label)

        self.scope_combo = QComboBox()
        self.scope_combo.addItem("All rows", "all")
        self.scope_combo.addItem("Current filters", "filters")
        self.scope_combo.addItem("Selected rows only", "selected")
        if not allow_selected:
            index = self.scope_combo.findData("selected")
            if index >= 0:
                self.scope_combo.model().item(index).setEnabled(False)
        layout.addWidget(self.scope_combo)

        format_label = QLabel("Format:")
        format_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(format_label)

        self.format_combo = QComboBox()
        self.format_combo.addItem("CSV", "csv")
        self.format_combo.addItem("XLSX", "xlsx")
        layout.addWidget(self.format_combo)

        self.ordering_checkbox = QCheckBox("Use current sort order")
        self.ordering_checkbox.setChecked(True)
        layout.addWidget(self.ordering_checkbox)

        fields_label = QLabel("Fields to include:")
        fields_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(fields_label)

        self.field_checks: list[QCheckBox] = []
        field_grid = QGridLayout()
        layout.addLayout(field_grid)
        for row, spec in enumerate(fields):
            check = QCheckBox(spec.label)
            check.setChecked(spec.in_export_default)
            check.setProperty("fieldKey", spec.key)
            field_grid.addWidget(check, row // 2, row % 2)
            self.field_checks.append(check)

        layout.addStretch(1)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_scope(self) -> str:
        return str(self.scope_combo.currentData())

    def selected_format(self) -> str:
        return str(self.format_combo.currentData())

    def selected_fields(self) -> list[str]:
        return [str(c.property("fieldKey")) for c in self.field_checks if c.isChecked()]

    def use_current_order(self) -> bool:
        return self.ordering_checkbox.isChecked()


def write_export_file(
    path: Path,
    rows: Sequence[dict[str, Any]],
    fields: Sequence[str],
    field_labels: dict[str, str],
    fmt: str,
) -> None:
    headers = [field_labels.get(f, f.title()) for f in fields]
    if fmt == "xlsx":
        from openpyxl import Workbook  # type: ignore

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append([_stringify(row.get(f)) for f in fields])
        workbook.save(path)
    else:
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([_stringify(row.get(f)) for f in fields])


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


# ---------------------------------------------------------------------------
# Async runner
# ---------------------------------------------------------------------------


class _AsyncWorker(QThread):
    done = Signal(object)
    failed = Signal(str)

    def __init__(self, fn: Callable[[], Any], parent: QObject | None = None) -> None:
        super().__init__(parent)
        self._fn = fn

    def run(self) -> None:  # type: ignore[override]
        try:
            result = self._fn()
        except Exception as exc:  # pragma: no cover - depends on runtime state
            self.failed.emit(str(exc))
        else:
            self.done.emit(result)


def run_async(
    owner: QObject,
    task: Callable[[], Any],
    on_done: Callable[[Any], None],
    on_error: Callable[[str], None],
) -> None:
    """Run ``task`` on a background ``QThread`` and deliver the result on the UI thread.

    ``owner`` must outlive the call (typically the window issuing the request); a
    reference to the worker is kept on ``owner._edit_kit_workers`` until it finishes
    so it isn't garbage-collected mid-run.
    """
    worker = _AsyncWorker(task, owner)

    workers: list[_AsyncWorker] = getattr(owner, "_edit_kit_workers", None)
    if workers is None:
        workers = []
        setattr(owner, "_edit_kit_workers", workers)
    workers.append(worker)

    def _cleanup(w: _AsyncWorker = worker) -> None:
        if w in workers:
            workers.remove(w)
        w.deleteLater()

    worker.done.connect(on_done)
    worker.failed.connect(on_error)
    worker.finished.connect(_cleanup)
    worker.start()
